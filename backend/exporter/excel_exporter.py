import io
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

_THIN = Side(style="thin", color="CCCCCC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_WARN_THIN   = Side(style="thin", color="FF4D6D")
_WARN_BORDER = Border(left=_WARN_THIN, right=_WARN_THIN, top=_WARN_THIN, bottom=_WARN_THIN)

_C = {
    "topic_bg":    "C6EFCE",
    "raw_bg":      "FFF2CC",
    "avro_bg":     "E2EFDA",
    "detail_bg":   "BDD7EE",
    "col_hdr_bg":  "2E75B6",
    "col_hdr_fg":  "FFFFFF",
    "pk_bg":       "FFD966",
    "logical_bg":  "FFFFC0",
    "row_odd":     "FFFFFF",
    "row_even":    "F2F2F2",
    # byte-anomaly warning colours
    "warn_bg":     "FFE7EC",   # light red fill
    "warn_fg":     "C0002A",   # dark red text
    "warn_hdr_bg": "FF4D6D",   # red header bar
    "warn_hdr_fg": "FFFFFF",
    "warn_border": "FF4D6D",
}


def _s(ws, row, col, value, bg=None, fg="000000", bold=False,
       align_h="center", wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    c.font = Font(name="Arial", size=10, bold=bold, color=fg)
    c.alignment = Alignment(horizontal=align_h, vertical="center", wrap_text=wrap)
    c.border = _BORDER
    return c


def _sw(ws, row, col, value, bold=False, align_h="left", wrap=True):
    """Write a warning-styled cell (red bg, red border)."""
    c = ws.cell(row=row, column=col, value=value)
    c.fill   = PatternFill("solid", start_color=_C["warn_bg"])
    c.font   = Font(name="Arial", size=10, bold=bold, color=_C["warn_fg"])
    c.alignment = Alignment(horizontal=align_h, vertical="center", wrap_text=wrap)
    c.border = _WARN_BORDER
    return c


def _write_warning_section(ws, anomalies: list, start_row: int, num_cols: int = 8) -> int:
    """
    เขียน WARNING section ต่อท้าย sheet สำหรับ byte anomaly
    anomalies = [{ column_name, source_type, raw_type, detail, file }, ...]
    """
    if not anomalies:
        return start_row

    last_col_letter = chr(ord('A') + num_cols - 1)   # 'H' เมื่อ num_cols=8

    # ── blank gap ─────────────────────────────────────────
    start_row += 1

    # ── header bar ────────────────────────────────────────
    ws.merge_cells(f"A{start_row}:{last_col_letter}{start_row}")
    c = ws.cell(row=start_row, column=1,
                value=f"⚠  WARNING — Byte Conversion Anomaly ({len(anomalies)} คอลัมน์)  ⚠")
    c.fill      = PatternFill("solid", start_color=_C["warn_hdr_bg"])
    c.font      = Font(name="Arial", size=11, bold=True, color=_C["warn_hdr_fg"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = _WARN_BORDER
    ws.row_dimensions[start_row].height = 20
    start_row += 1

    # ── sub-header ────────────────────────────────────────
    ws.merge_cells(f"A{start_row}:{last_col_letter}{start_row}")
    _sw(ws, start_row, 1,
        "คอลัมน์ด้านล่างถูกแปลงเป็น byte แต่ type ต้นทางไม่ใช่ decimal-family "
        "— กรุณาตรวจสอบ mapping และแก้ไขจุดที่ผิดพลาด",
        bold=False, wrap=True)
    ws.row_dimensions[start_row].height = 30
    start_row += 1

    # ── column headers ────────────────────────────────────
    headers = ["NO.", "Column Name", "Source SQL Type", "Raw Type (byte)", "Detail / คำอธิบาย", "File"]
    header_widths = [6, 22, 18, 18, 52, 28]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=start_row, column=ci, value=h)
        c.fill      = PatternFill("solid", start_color=_C["warn_hdr_bg"])
        c.font      = Font(name="Arial", size=10, bold=True, color=_C["warn_hdr_fg"])
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _WARN_BORDER
    # merge ส่วนที่เกินหัวตาราง (num_cols อาจเกิน 6 col)
    if num_cols > len(headers):
        ws.merge_cells(
            start_row=start_row, start_column=len(headers),
            end_row=start_row,   end_column=num_cols
        )
    start_row += 1

    # ── data rows ─────────────────────────────────────────
    for i, a in enumerate(anomalies, 1):
        _sw(ws, start_row, 1, i,              align_h="center", wrap=False)
        _sw(ws, start_row, 2, a.get("column_name", ""))
        _sw(ws, start_row, 3, a.get("source_type", ""))
        _sw(ws, start_row, 4, a.get("raw_type", ""))
        _sw(ws, start_row, 5, a.get("detail",  ""), wrap=True)
        _sw(ws, start_row, 6, a.get("file",    ""))
        # fill ช่องที่เกิน (ถ้า num_cols > 6)
        for ci in range(7, num_cols + 1):
            _sw(ws, start_row, ci, "")
        ws.row_dimensions[start_row].height = 28
        start_row += 1

    # ปรับความกว้างคอลัมน์ warning (ใช้เฉพาะถ้า col ยังไม่ได้ set)
    for ci, w in enumerate(header_widths, 1):
        col_letter = chr(ord('A') + ci - 1)
        if ws.column_dimensions[col_letter].width < w:
            ws.column_dimensions[col_letter].width = w

    return start_row


def _write_raw_section(ws, table_name: str, columns: list, start_row: int) -> int:
    """
    Section 1 — Raw (SQL Server)
    NO. | Name | PK or Unique | Max Length | Format | Nullable | Description | Possible Value
    """
    # TABLE header
    ws.merge_cells(f"A{start_row}:H{start_row}")
    _s(ws, start_row, 1, f"TABLE:    {table_name}",
       bg=_C["topic_bg"], bold=True, align_h="left")
    start_row += 1

    # Raw (SQL Server) label
    ws.merge_cells(f"A{start_row}:H{start_row}")
    _s(ws, start_row, 1, "Raw (SQL Server)",
       bg=_C["raw_bg"], bold=True, align_h="left")
    start_row += 1

    # Detail Section
    ws.merge_cells(f"A{start_row}:H{start_row}")
    _s(ws, start_row, 1, "Detail Section",
       bg=_C["detail_bg"], bold=True, align_h="left")
    start_row += 1

    # Column headers
    for c, h in enumerate(
        ["NO.", "Name", "PK or Unique", "Max Length",
         "Format", "Nullable", "Description", "Possible Value"], 1
    ):
        _s(ws, start_row, c, h,
           bg=_C["col_hdr_bg"], fg=_C["col_hdr_fg"], bold=True)
    start_row += 1

    # Data rows
    for i, col in enumerate(columns, 1):
        r = start_row
        bg = _C["row_odd"] if i % 2 == 1 else _C["row_even"]
        is_pk = "Y" if col.get("is_pk") else "N"

        sql_type = col.get("source_sql_type", "")
        m = re.search(r"\(([^)]+)\)", sql_type)
        max_len = m.group(1) if m else "-"

        base_type = re.split(r"[\(\s]", sql_type.lower().strip())[0]

        _s(ws, r, 1, i,           bg=bg)
        _s(ws, r, 2, col.get("column_name", ""),   bg=bg, align_h="left")
        _s(ws, r, 3, is_pk,       bg=_C["pk_bg"] if is_pk == "Y" else bg)
        _s(ws, r, 4, max_len,     bg=bg)
        _s(ws, r, 5, base_type,   bg=bg)
        _s(ws, r, 6, col.get("nullable", ""),      bg=bg)
        _s(ws, r, 7, "",          bg=bg, align_h="left", wrap=True)
        _s(ws, r, 8, "",          bg=bg, align_h="left", wrap=True)
        start_row += 1

    return start_row + 1  # blank gap


def _write_avro_section(ws, table_name: str, columns: list, start_row: int) -> int:
    """
    Section 2 — Confluent (AVRO)
    NO. | Name | Partition Key | Raw Format type | Logical Format type | direct move / logic | Description | Possible Value
    """
    # Topic row
    ws.merge_cells(f"A{start_row}:H{start_row}")
    _s(ws, start_row, 1,
       f"Topic:    UAT_EEAS_RAW_dbWorkforce_{table_name}",
       bg=_C["topic_bg"], bold=True, align_h="left")
    start_row += 1

    # Confluent (AVRO) label
    ws.merge_cells(f"A{start_row}:H{start_row}")
    _s(ws, start_row, 1, "Confluent (AVRO)",
       bg=_C["avro_bg"], bold=True, align_h="left")
    start_row += 1

    # Detail Section
    ws.merge_cells(f"A{start_row}:H{start_row}")
    _s(ws, start_row, 1, "Detail Section",
       bg=_C["detail_bg"], bold=True, align_h="left")
    start_row += 1

    # Column headers
    for c, h in enumerate(
        ["NO.", "Name", "Partition Key", "Raw Format type",
         "Logical Format type", "direct move / logic", "Description", "Possible Value"], 1
    ):
        _s(ws, start_row, c, h,
           bg=_C["col_hdr_bg"], fg=_C["col_hdr_fg"], bold=True)
    start_row += 1

    # Data rows
    for i, col in enumerate(columns, 1):
        r = start_row
        bg = _C["row_odd"] if i % 2 == 1 else _C["row_even"]

        _s(ws, r, 1, i,                              bg=bg)
        _s(ws, r, 2, col.get("column_name", ""),     bg=bg, align_h="left")
        _s(ws, r, 3, "",                             bg=bg)   # Partition Key — ให้ user กรอกเอง
        _s(ws, r, 4, col.get("raw_type", ""),        bg=bg)
        _s(ws, r, 5, col.get("logical_type", ""),    bg=_C["logical_bg"])
        _s(ws, r, 6, "Direct move",                  bg=bg)
        _s(ws, r, 7, "",                             bg=bg, align_h="left", wrap=True)  # Description
        _s(ws, r, 8, "",                             bg=bg, align_h="left", wrap=True)  # Possible Value
        start_row += 1

    return start_row + 1


def _build_sheet(ws, table_name: str, columns: list, anomalies: list | None = None):
    """1 sheet = Raw section + AVRO section + (optional) WARNING section"""
    next_row = _write_raw_section(ws, table_name, columns, start_row=1)
    next_row = _write_avro_section(ws, table_name, columns, start_row=next_row)

    if anomalies:
        _write_warning_section(ws, anomalies, start_row=next_row, num_cols=8)

    for col, w in zip("ABCDEFGH", [8, 22, 14, 17, 19, 18, 40, 40]):
        ws.column_dimensions[col].width = w


# ── Public API ────────────────────────────────────────────────────

def export_confluent_xlsx(tables: dict, byte_anomalies: dict | None = None) -> io.BytesIO:
    wb = Workbook()
    wb.remove(wb.active)
    for table_name, columns in tables.items():
        ws = wb.create_sheet(title=table_name[:31])
        anomalies = (byte_anomalies or {}).get(table_name)
        _build_sheet(ws, table_name, columns, anomalies=anomalies)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_table_xlsx(columns: list, table_name: str = "Sheet1",
                      anomalies: list | None = None) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = table_name[:31]
    _build_sheet(ws, table_name, columns, anomalies=anomalies)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ── Legacy (ชื่อเดิม — main.py ไม่ต้องแก้) ───────────────────────

def export_all_xlsx(tables: dict) -> io.BytesIO:
    return export_confluent_xlsx(tables)


def _build_csv_rows(table_name: str, columns: list, anomalies: list | None = None) -> list:
    import re
    rows = []

    # Raw section
    rows.append([f"TABLE:    {table_name}"])
    rows.append(["Raw (SQL Server)"])
    rows.append(["Detail Section"])
    rows.append(["NO.", "Name", "PK or Unique", "Max Length", "Format", "Nullable", "Description", "Possible Value"])
    for i, col in enumerate(columns, 1):
        sql_type = col.get("source_sql_type", "")
        m = re.search(r"\(([^)]+)\)", sql_type)
        max_len = m.group(1) if m else "-"
        base_type = re.split(r"[\(\s]", sql_type.lower().strip())[0]
        rows.append([i, col.get("column_name", ""), "Y" if col.get("is_pk") else "N",
                     max_len, base_type, col.get("nullable", ""), "", ""])

    rows.append([])  # blank gap

    # Confluent section
    rows.append([f"Topic:    UAT_EEAS_RAW_dbWorkforce_{table_name}"])
    rows.append(["Confluent (AVRO)"])
    rows.append(["Detail Section"])
    rows.append(["NO.", "Name", "Partition Key", "Raw Format type", "Logical Format type", "direct move / logic", "Description", "Possible Value"])
    for i, col in enumerate(columns, 1):
        rows.append([i, col.get("column_name", ""), "",   # Partition Key — ให้ user กรอกเอง
                     col.get("raw_type", ""), col.get("logical_type", ""), "Direct move", "", ""])

    # ── WARNING section (byte anomaly) ───────────────────
    if anomalies:
        rows.append([])
        rows.append([f"⚠ WARNING — Byte Conversion Anomaly ({len(anomalies)} คอลัมน์) ⚠"])
        rows.append(["คอลัมน์ด้านล่างถูกแปลงเป็น byte แต่ type ต้นทางไม่ใช่ decimal-family "
                     "— กรุณาตรวจสอบ mapping และแก้ไขจุดที่ผิดพลาด"])
        rows.append(["NO.", "Column Name", "Source SQL Type", "Raw Type (byte)", "Detail / คำอธิบาย", "File"])
        for i, a in enumerate(anomalies, 1):
            rows.append([
                i,
                a.get("column_name", ""),
                a.get("source_type",  ""),
                a.get("raw_type",     ""),
                a.get("detail",       ""),
                a.get("file",         ""),
            ])

    return rows


def _csv_bytes(rows: list) -> io.BytesIO:
    import csv
    import codecs
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerows(rows)
    output = io.BytesIO(codecs.BOM_UTF8 + buf.getvalue().encode("utf-8"))
    output.seek(0)
    return output


def export_all_csv(tables: dict, byte_anomalies: dict | None = None) -> io.BytesIO:
    all_rows = []
    first = True
    for table_name, columns in tables.items():
        if not first:
            all_rows.append([])
        anomalies = (byte_anomalies or {}).get(table_name)
        all_rows.extend(_build_csv_rows(table_name, columns, anomalies=anomalies))
        first = False
    return _csv_bytes(all_rows)


def export_table_csv(columns: list, table_name: str = "Sheet1",
                     anomalies: list | None = None) -> io.BytesIO:
    return _csv_bytes(_build_csv_rows(table_name, columns, anomalies=anomalies))